"""
Australia — CRICOS.

The flagship source, and the reason Australia is built first: data.gov.au keeps
every dated edition, so the history does not have to be accumulated going
forward. It can be reconstructed backwards on day one.

Verified 25 August 2026:
  * 57 dated editions, unbroken monthly, 2021-10-31 to 2026-07-01
  * historical editions are XLSX workbooks, NOT the four CSVs the source's
    current-edition files might lead you to expect
  * five sheets: Purpose Statement, Institutions, Courses, Locations,
    Course Locations
  * row 0 is the sheet title, row 1 carries the publisher's own edition stamp
    ("Report generated Wednesday, 1 July 2026"), row 2 is the header,
    data starts at row 3
  * Institutions key on CRICOS Provider Code, Courses on provider + course code

Measured churn, July 2025 to July 2026:
    providers   1,529 -> 1,545     86 removed, 102 added, 18 renamed
    courses    25,415 -> 26,604  1,666 removed, 2,855 added
    of those removed courses, 1,114 sat at a provider that is STILL LISTED
"""

from __future__ import annotations

import datetime as _dt
import io
import re
from typing import Any, Iterable

import requests
from openpyxl import load_workbook

CKAN_PACKAGE = "https://data.gov.au/data/api/3/action/package_show?id=cricos"
USER_AGENT = (
    "studiesmultiverse-standing/1.0 (+https://studiesmultiverse.com/standing/methodology/) "
    "public-register archival bot"
)

INSTITUTION_KEY = "CRICOS Provider Code"
INSTITUTION_NAME = "Institution Name"
COURSE_KEY_FIELDS = ("CRICOS Provider Code", "CRICOS Course Code")
COURSE_NAME = "Course Name"

# Fields worth watching for change at provider level. Quoted verbatim, never
# paraphrased into a verdict.
PROVIDER_WATCH_FIELDS = ("Institution Type", "Institution Capacity", "Postal Address State")

_EDITION_RE = re.compile(r"^\s*(20\d\d-\d\d-\d\d)")
_GENERATED_RE = re.compile(r"generated\s+(?:\w+day,?\s*)?(\d{1,2}\s+\w+\s+20\d\d)", re.I)


def _session() -> requests.Session:
    s = requests.Session()
    s.headers.update({"User-Agent": USER_AGENT})
    return s


def list_editions(session: requests.Session | None = None) -> list[dict]:
    """
    Every dated edition on the dataset, oldest first.

    Returns dicts of {edition_date, name, url, format, size}.
    """
    s = session or _session()
    payload = s.get(CKAN_PACKAGE, timeout=60).json()
    out = []
    for res in payload["result"]["resources"]:
        m = _EDITION_RE.match(res.get("name", ""))
        if not m:
            continue
        out.append(
            {
                "edition_date": m.group(1),
                "name": res["name"].strip(),
                "url": res["url"],
                "format": (res.get("format") or "").upper(),
                "size": res.get("size"),
            }
        )
    out.sort(key=lambda r: r["edition_date"])
    return out


def fetch_edition(url: str, session: requests.Session | None = None) -> bytes:
    s = session or _session()
    r = s.get(url, timeout=300)
    r.raise_for_status()
    return r.content


def _sheet_rows(ws) -> list[list[Any]]:
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _parse_sheet(ws) -> tuple[list[dict], str | None]:
    """
    Returns (rows, publisher_edition_stamp).

    Row 0 title, row 1 the stamp, row 2 the header, data from row 3.
    Falls back to locating the header row if the publisher changes the layout —
    but a failure here surfaces as missing required columns at the sanity gate
    rather than being silently papered over.
    """
    raw = _sheet_rows(ws)
    if len(raw) < 4:
        return [], None

    stamp = None
    for cell in (raw[1] or [])[:3]:
        if cell and isinstance(cell, str) and "generated" in cell.lower():
            stamp = cell.strip()
            break

    header_idx = 2
    if not any(isinstance(c, str) and c.strip() == INSTITUTION_KEY for c in (raw[2] or [])):
        for i, row in enumerate(raw[:8]):
            if any(isinstance(c, str) and c.strip() in (INSTITUTION_KEY, "CRICOS Course Code") for c in (row or [])):
                header_idx = i
                break

    header = [str(c).strip() if c is not None else "" for c in raw[header_idx]]
    rows: list[dict] = []
    for values in raw[header_idx + 1 :]:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        rec = {}
        for i, col in enumerate(header):
            if not col:
                continue
            v = values[i] if i < len(values) else None
            rec[col] = "" if v is None else str(v).strip()
        if rec:
            rows.append(rec)
    return rows, stamp


def parse_edition(data: bytes) -> dict:
    """
    Parse one CRICOS workbook.

    Returns {source_date, institutions, courses, locations_count, course_locations_count}.
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        institutions, stamp = _parse_sheet(wb["Institutions"]) if "Institutions" in wb.sheetnames else ([], None)
        courses, stamp2 = _parse_sheet(wb["Courses"]) if "Courses" in wb.sheetnames else ([], None)

        counts = {}
        for extra in ("Locations", "Course Locations"):
            if extra in wb.sheetnames:
                counts[extra] = max(0, wb[extra].max_row - 3)

        return {
            "source_date": _stamp_to_date(stamp or stamp2),
            "source_stamp": stamp or stamp2,
            "institutions": institutions,
            "courses": courses,
            "sheet_counts": counts,
        }
    finally:
        wb.close()


def _stamp_to_date(stamp: str | None) -> str | None:
    """'Report generated Wednesday, 1 July 2026' -> '2026-07-01'."""
    if not stamp:
        return None
    m = _GENERATED_RE.search(stamp)
    if not m:
        return None
    for fmt in ("%d %B %Y", "%d %b %Y"):
        try:
            return _dt.datetime.strptime(m.group(1), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def course_key(row: dict) -> str:
    return "|".join(str(row.get(f, "")).strip() for f in COURSE_KEY_FIELDS)


def add_course_keys(courses: Iterable[dict]) -> list[dict]:
    out = []
    for r in courses:
        r = dict(r)
        r["_key"] = course_key(r)
        out.append(r)
    return out


def courses_removed_at_live_providers(
    old_courses: list[dict], new_courses: list[dict], new_institutions: list[dict]
) -> list[dict]:
    """
    The finding that justifies course-level tracking at all.

    A course that disappears while its provider remains listed is invisible to
    every provider-level monitor — and it is the case that actually affects the
    student, because the student is enrolled on a course, not on a provider.

    Measured over twelve months to July 2026: 1,114 of 1,666 removed courses
    sat at a provider still on the register.
    """
    live = {str(r.get(INSTITUTION_KEY, "")).strip() for r in new_institutions}
    new_keys = {course_key(r) for r in new_courses}
    out = []
    for r in old_courses:
        k = course_key(r)
        if k in new_keys:
            continue
        provider = str(r.get(INSTITUTION_KEY, "")).strip()
        if provider in live:
            out.append(
                {
                    "provider_code": provider,
                    "provider_name": r.get(INSTITUTION_NAME, ""),
                    "course_code": r.get("CRICOS Course Code", ""),
                    "course_name": r.get(COURSE_NAME, ""),
                    "course_level": r.get("Course Level", ""),
                }
            )
    return out
